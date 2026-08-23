#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_JSON = REPO_ROOT / "calendar-data-2026-2027.json"
OUTPUT_XLSX = REPO_ROOT / "kalendarz_dzienny_role_2026_2027.xlsx"

START_DAY = date(2026, 9, 1)
END_DAY = date(2027, 8, 31)

HEADERS = [
    "data",
    "dzień tygodnia",
    "opis 1 (wolne/święta/inne)",
    "opis 2: wychowawca",
    "opis 3: nauczyciel",
    "opis 4: dyrektor",
]

WEEKDAY_NAMES = {
    0: "poniedziałek",
    1: "wtorek",
    2: "środa",
    3: "czwartek",
    4: "piątek",
    5: "sobota",
    6: "niedziela",
}


def load_events() -> list[dict]:
    payload = json.loads(SOURCE_JSON.read_text(encoding="utf-8"))
    return payload["events"]


def parse_day(value: str) -> date:
    return datetime.fromisoformat(value[:10]).date()


def event_days(event: dict) -> list[date]:
    start_day = parse_day(event["start"])
    if "end" not in event:
        return [start_day]

    end_day = parse_day(event["end"])
    if event.get("allDay", False):
        end_day -= timedelta(days=1)

    if end_day < start_day:
        return [start_day]

    days: list[date] = []
    current = start_day
    while current <= end_day:
        days.append(current)
        current += timedelta(days=1)
    return days


def compact_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"\n\s*\n+", "\n", value)
    value = re.sub(r"[ \t]+", " ", value)
    return value.strip()


def event_label(event: dict) -> str:
    description = compact_text(event.get("extendedProps", {}).get("description", "")).strip()
    title = compact_text(event.get("title", "")).strip()
    base = (description or title).replace("\n", " / ")
    if "T" in event["start"]:
        time_part = event["start"].split("T", 1)[1][:5]
        return f"{time_part} - {base}"
    return base


def role_column(event: dict) -> int:
    text = "\n".join(
        [
            event.get("title", ""),
            event.get("extendedProps", {}).get("description", ""),
            event.get("extendedProps", {}).get("audience", ""),
        ]
    ).lower()

    if any(token in text for token in ["dyrektor", "dyrekcj", "wicedyrektor"]):
        return 5

    if re.search(r"\bwychowawc(a|y|ów|om|ami|ach)?\b", text) or any(
        token in text for token in ["rodziców", "rodzice"]
    ):
        return 3

    if (
        event.get("extendedProps", {}).get("audience") == "nauczyciele"
        or any(
            token in text
            for token in [
                "rada pedagogiczna",
                "rada plenarna",
                "klasyfikacyjna",
                "inauguracyjna",
                "szkolenie",
                "egzamin",
                "matura",
                "nadzoru pedagogicznego",
            ]
        )
    ):
        return 4

    return 2


def build_rows(events: list[dict]) -> dict[date, dict[int, list[str]]]:
    rows = {
        START_DAY + timedelta(days=offset): {2: [], 3: [], 4: [], 5: []}
        for offset in range((END_DAY - START_DAY).days + 1)
    }

    for event in events:
        label = event_label(event)
        target_column = role_column(event)
        for day in event_days(event):
            if day not in rows:
                continue
            rows[day][target_column].append(label)
    return rows


def autosize(ws) -> None:
    widths = {
        "A": 14,
        "B": 18,
        "C": 42,
        "D": 42,
        "E": 42,
        "F": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DFF5E7")
    border = Border(
        left=Side(style="thin", color="D5E4DA"),
        right=Side(style="thin", color="D5E4DA"),
        top=Side(style="thin", color="D5E4DA"),
        bottom=Side(style="thin", color="D5E4DA"),
    )

    for cell in ws[1]:
        cell.font = Font(bold=True, color="16332B")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"


def main() -> None:
    events = load_events()
    day_rows = build_rows(events)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kalendarz dzienny"
    ws.append(HEADERS)

    current = START_DAY
    while current <= END_DAY:
        row_map = day_rows[current]
        ws.append(
            [
                current,
                WEEKDAY_NAMES[current.weekday()],
                "\n\n".join(row_map[2]),
                "\n\n".join(row_map[3]),
                "\n\n".join(row_map[4]),
                "\n\n".join(row_map[5]),
            ]
        )
        current += timedelta(days=1)

    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    autosize(ws)
    style_sheet(ws)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
