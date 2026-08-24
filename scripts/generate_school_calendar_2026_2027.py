#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = REPO_ROOT / "kalendarz_dzienny_role_2026_2027.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "calendar-data-2026-2027.json"

ROLE_COLUMNS = ["opis1", "opis2", "opis3", "opis4"]
ROLE_AUDIENCE = {
    "opis1": "wszyscy",
    "opis2": "wychowawcy",
    "opis3": "nauczyciele",
    "opis4": "dyrekcja",
}
PLACEHOLDER_LINES = {
    "Zakres wygenerowany z dziennych wpisów XLSX.",
    "Zakres zgodnie z XLSX: 5.10.2026 - 30.10.2026.",
    "Zakres zgodnie z XLSX: 2.02.2027 - 26.02.2027.",
}
KNOWN_RANGE_TITLES = {
    ("2026-10-05", "2026-10-30"): "Praktyki zawodowe - klasy 3 technikum",
    ("2026-12-23", "2027-01-03"): "Zimowa przerwa świąteczna",
    ("2027-01-18", "2027-01-31"): "Ferie zimowe",
    ("2027-02-02", "2027-02-26"): "Praktyki zawodowe - klasy 4 technikum",
    ("2027-04-01", "2027-04-06"): "Wiosenna przerwa świąteczna",
}
SERIES_LINE_PREFIXES = (
    "Zimowa przerwa świąteczna",
    "Wiosenna przerwa świąteczna",
    "Ferie zimowe",
    "KLASY 3 TECHNIKUM: praktyki",
    "KLASY 4 TECHNIKUM: praktyki",
    "Praktyki zawodowe - klasy 3 technikum",
    "Praktyki zawodowe - klasy 4 technikum",
)


@dataclass
class EventDraft:
    start: str
    end: str | None
    all_day: bool
    title: str
    description: str
    category: str
    tags: list[str]
    priority: str
    audience: str
    needs_confirmation: bool = False
    source_note: str | None = None

    def to_dict(self, idx: int) -> dict:
        payload = {
            "id": f"evt-{idx:04d}",
            "title": self.title,
            "start": self.start,
            "allDay": self.all_day,
            "category": self.category,
            "tags": self.tags,
            "priority": self.priority,
            "audience": self.audience,
            "extendedProps": {
                "description": self.description,
                "needsConfirmation": self.needs_confirmation,
                "sourceNote": self.source_note or "",
            },
        }
        if self.end:
            payload["end"] = self.end
        return payload


def compact(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slug(text: str) -> str:
    lowered = text.lower()
    lowered = (
        lowered.replace("ą", "a")
        .replace("ć", "c")
        .replace("ę", "e")
        .replace("ł", "l")
        .replace("ń", "n")
        .replace("ó", "o")
        .replace("ś", "s")
        .replace("ź", "z")
        .replace("ż", "z")
    )
    return re.sub(r"[^a-z0-9]+", "-", lowered).strip("-")


def split_chunks(text: str) -> list[str]:
    if not text:
        return []
    normalized = compact(text)
    if not normalized:
        return []
    chunks: list[str] = []
    for chunk in re.split(r"\n\s*\n", normalized):
        stripped_chunk = chunk.strip()
        if not stripped_chunk:
            continue
        lines = [line.strip() for line in stripped_chunk.split("\n") if line.strip()]
        if (
            len(lines) >= 2
            and "dzień ustawowo wolny" in lines[0].lower()
            and any(lines[1].startswith(prefix) for prefix in SERIES_LINE_PREFIXES)
        ):
            chunks.append(lines[0])
            chunks.append("\n".join(lines[1:]))
            continue
        chunks.append(stripped_chunk)
    return chunks


def normalize_chunk_text(text: str) -> str:
    return compact(text).replace(" / ", "\n")


def strip_leading_time_prefix(text: str) -> str:
    cleaned = re.sub(
        r"^\s*\d{1,2}:\d{2}(?:\s*[/-]\s*\d{1,2}:\d{2})?\s*[-–:]\s*",
        "",
        text.strip(),
    )
    return cleaned or text.strip()


def short_title(text: str, fallback: str) -> str:
    normalized = normalize_chunk_text(text)
    first_line = normalized.split("\n")[0].strip(" -;")
    title = strip_leading_time_prefix(first_line) or fallback
    if len(title) > 88:
        title = title[:85].rstrip() + "..."
    return title


def parse_time(text: str) -> str | None:
    match = re.search(r"(\d{1,2}:\d{2})", text)
    if not match:
        return None
    hour, minute = match.group(1).split(":")
    return f"{int(hour):02d}:{minute}"


def parse_polish_day(text: str) -> date | None:
    match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if not match:
        return None
    day, month, year = [int(part) for part in match.groups()]
    return date(year, month, day)


def parse_range_dates(text: str) -> tuple[date, date] | None:
    matches = re.findall(r"(\d{1,2}\.\d{1,2}\.\d{4})", text)
    if len(matches) < 2:
        return None
    start = parse_polish_day(matches[0])
    end = parse_polish_day(matches[1])
    if not start or not end:
        return None
    return start, end


def detect_category(text: str) -> tuple[str, list[str], str]:
    lower = text.lower()
    tags: set[str] = set()
    category = "general"
    priority = "normal"

    if "rada pedagogiczna" in lower or "rada " in lower:
        category = "council"
        tags.add("rady")
        priority = "high"
    if any(
        token in lower
        for token in [
            "wystawienie ocen",
            "zestawień klasyfikacyjnych",
            "klasyfikacyjna",
            "proponowanych ocen",
            "ocen rocznych",
            "ocen końcowych",
        ]
    ):
        category = "classification"
        tags.add("klasyfikacja")
        priority = "high"
    if any(token in lower for token in ["matura", "egzamin", "egzaminy", "egzaminu semestralnego", "egzamin zawodowy"]):
        if category == "general":
            category = "exam"
        tags.add("egzaminy")
        priority = "high"
    if "praktyki" in lower:
        if category == "general":
            category = "practice"
        tags.add("praktyki")
    if any(token in lower for token in ["zebranie", "zebrania", "konsultacje", "rada rodziców"]):
        if category == "general":
            category = "meeting"
        tags.add("spotkania")
    if any(token in lower for token in ["wolne", "ferie", "przerwa świąteczna", "święto", "boże ciało", "wielkanoc"]):
        if category == "general":
            category = "holiday"
        tags.add("wolne")
    if "bs ii" in lower or "bs2st" in lower or "bs ii st." in lower:
        tags.add("bs2")
    if any(token in lower for token in ["wymagający wyjaśnienia", "wymaga doprecyzowania", "wymagający doprecyzowania"]):
        tags.add("doprecyzowanie")
        priority = "important"
    if any(token in lower for token in ["jubileusz", "święto szkoły", "rozpoczęcie roku", "zakończenie roku"]):
        priority = "high"
        tags.add("kluczowe")

    return category, sorted(tags), priority


def detect_audience(text: str, role_key: str) -> str:
    lower = text.lower()
    if role_key == "opis4":
        return "dyrekcja"
    if role_key == "opis3":
        return "nauczyciele"
    if role_key == "opis2":
        return "rodzice i wychowawcy" if "rodzic" in lower else "wychowawcy"
    if "5 technikum" in lower or "klasy 5" in lower:
        return "klasy maturalne"
    if "bs ii" in lower or "bs2st" in lower:
        return "bs ii stopnia"
    if "praktyki" in lower:
        return "wybrane klasy"
    return ROLE_AUDIENCE[role_key]


def rows_from_sheet(source_path: Path) -> list[dict]:
    wb = load_workbook(source_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [compact(str(value)).lower() if value not in (None, "") else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def find_index(*candidates: str, default: int) -> int:
        for idx, header in enumerate(headers):
            for candidate in candidates:
                if header.startswith(candidate):
                    return idx
        return default

    date_idx = find_index("data", default=0)
    opis1_idx = find_index("opis 1", "opis1", default=2)
    opis2_idx = find_index("opis 2", "opis2", default=3)
    opis3_idx = find_index("opis 3", "opis3", default=4)
    opis4_idx = find_index("opis 4", "opis4", default=5)

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        dt = values[date_idx]
        if not dt:
            continue
        rows.append(
            {
                "date": dt.date() if isinstance(dt, datetime) else dt,
                "opis1": compact(str(values[opis1_idx])) if values[opis1_idx] not in (None, "") else "",
                "opis2": compact(str(values[opis2_idx])) if values[opis2_idx] not in (None, "") else "",
                "opis3": compact(str(values[opis3_idx])) if values[opis3_idx] not in (None, "") else "",
                "opis4": compact(str(values[opis4_idx])) if values[opis4_idx] not in (None, "") else "",
            }
        )
    return rows


def is_placeholder_only(chunk: str) -> bool:
    return compact(chunk) in PLACEHOLDER_LINES


def is_range_chunk(chunk: str) -> bool:
    normalized = normalize_chunk_text(chunk)
    return (
        "Zakres wygenerowany z dziennych wpisów XLSX." in normalized
        or "Zakres zgodnie z XLSX:" in normalized
        or normalized.startswith("Zimowa przerwa świąteczna - ")
        or normalized.startswith("Wiosenna przerwa świąteczna - ")
        or normalized.startswith("KLASY 3 TECHNIKUM: praktyki")
        or normalized.startswith("KLASY 4 TECHNIKUM: praktyki")
    )


def next_day_iso(day: date) -> str:
    return (day + timedelta(days=1)).strftime("%Y-%m-%d")


def parse_iso_day(value: str) -> date:
    return date.fromisoformat(value[:10])


def event_flags(day: date, text: str) -> tuple[bool, str | None]:
    lower = text.lower()
    if day == date(2027, 6, 17) and "3 semestr bs ii" in lower:
        return True, "Wpis BS II przy 17.06.2027 pozostaje do doprecyzowania zgodnie z uwagą użytkownika."
    if day == date(2027, 6, 28) and "wymagający wyjaśnienia" in lower:
        return True, "Wpis z 28.06.2027 pozostaje oznaczony do wyjaśnienia zgodnie z treścią arkusza."
    return False, None


def build_single_day_event(day: date, chunk: str, role_key: str) -> EventDraft:
    description = normalize_chunk_text(chunk)
    category, tags, priority = detect_category(description)
    audience = detect_audience(description, role_key)
    title = short_title(description, "Wydarzenie")
    needs_confirmation, source_note = event_flags(day, description)
    if day == date(2027, 6, 28) and "Arkusz XLSX zawiera pod datą 28.06.2027" in description:
        title = "Rada pedagogiczna - wpis wymaga doprecyzowania"
    time_text = parse_time(description)

    if (
        time_text
        and "do ustalenia" not in description.lower()
        and not (day == date(2027, 6, 28) and needs_confirmation)
    ):
        start = f"{day.strftime('%Y-%m-%d')}T{time_text}:00"
        hour, minute = [int(part) for part in time_text.split(":")]
        end = (
            datetime(day.year, day.month, day.day, hour, minute) + timedelta(minutes=90)
        ).strftime("%Y-%m-%dT%H:%M:%S")
        all_day = False
    else:
        start = day.strftime("%Y-%m-%d")
        end = None
        all_day = True

    return EventDraft(
        start=start,
        end=end,
        all_day=all_day,
        title=title,
        description=description,
        category=category,
        tags=tags,
        priority=priority,
        audience=audience,
        needs_confirmation=needs_confirmation,
        source_note=source_note,
    )


def build_range_event(chunk: str, role_key: str) -> EventDraft | None:
    normalized = normalize_chunk_text(chunk)
    parsed_range = parse_range_dates(normalized)
    if not parsed_range:
        return None
    start_day, end_day = parsed_range
    preferred_title = KNOWN_RANGE_TITLES.get(
        (start_day.strftime("%Y-%m-%d"), end_day.strftime("%Y-%m-%d"))
    )
    title_lines = []
    for line in normalized.split("\n"):
        stripped = line.strip()
        if not stripped or stripped in PLACEHOLDER_LINES or stripped.startswith("Zakres zgodnie z XLSX:"):
            continue
        title_lines.append(stripped)

    title = preferred_title or title_lines[0] if title_lines else preferred_title or "Wydarzenie wielodniowe"
    description = "\n".join(title_lines) if title_lines else title
    category, tags, priority = detect_category(title)
    audience = detect_audience(title, role_key)
    if preferred_title and not title_lines:
        description = title

    return EventDraft(
        start=start_day.strftime("%Y-%m-%d"),
        end=next_day_iso(end_day),
        all_day=True,
        title=title,
        description=description,
        category=category,
        tags=tags,
        priority=priority,
        audience=audience,
    )


def build_events(rows: list[dict]) -> list[EventDraft]:
    events: list[EventDraft] = []
    seen_range_keys: set[tuple[str, str, str, str]] = set()

    for row in rows:
        day = row["date"]
        for role_key in ROLE_COLUMNS:
            for chunk in split_chunks(row[role_key]):
                if is_placeholder_only(chunk):
                    continue
                if is_range_chunk(chunk):
                    range_event = build_range_event(chunk, role_key)
                    if range_event:
                        range_key = (
                            role_key,
                            range_event.start,
                            range_event.end or "",
                        )
                        if range_key not in seen_range_keys:
                            seen_range_keys.add(range_key)
                            events.append(range_event)
                        continue
                events.append(build_single_day_event(day, chunk, role_key))

    return events


def normalize_series_text(text: str) -> str:
    normalized_lines: list[str] = []
    for raw_line in compact(text).split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"\s*\(dzień\s+\d+\s+z\s+\d+\)", "", line, flags=re.IGNORECASE)
        line = re.sub(r"\s*·\s*weekend\s*$", "", line, flags=re.IGNORECASE)
        if line.startswith("Dodatkowo:"):
            continue
        line = re.sub(r"\s{2,}", " ", line).strip()
        if line:
            normalized_lines.append(line)
    return "\n".join(normalized_lines)


def merge_signature(event: EventDraft) -> tuple:
    return (
        normalize_series_text(event.title),
        normalize_series_text(event.description),
        event.category,
        tuple(event.tags),
        event.priority,
        event.audience,
        event.needs_confirmation,
        event.source_note or "",
    )


def is_mergeable_single_day_event(event: EventDraft) -> bool:
    return event.all_day and not event.end and not event.needs_confirmation


def collapse_run(run: list[EventDraft]) -> list[EventDraft]:
    if len(run) == 1:
        return run
    first = run[0]
    last = run[-1]
    return [
        EventDraft(
            start=first.start,
            end=next_day_iso(parse_iso_day(last.start)),
            all_day=True,
            title=normalize_series_text(first.title),
            description=normalize_series_text(first.description),
            category=first.category,
            tags=first.tags,
            priority=first.priority,
            audience=first.audience,
            needs_confirmation=first.needs_confirmation,
            source_note=first.source_note,
        )
    ]


def merge_consecutive_single_day_events(events: list[EventDraft]) -> list[EventDraft]:
    grouped: dict[tuple, list[EventDraft]] = defaultdict(list)
    passthrough: list[EventDraft] = []

    for event in events:
        if is_mergeable_single_day_event(event):
            grouped[merge_signature(event)].append(event)
        else:
            passthrough.append(event)

    merged: list[EventDraft] = []
    for group_events in grouped.values():
        ordered = sorted(group_events, key=lambda event: event.start)
        run: list[EventDraft] = [ordered[0]]
        for event in ordered[1:]:
            previous_day = parse_iso_day(run[-1].start)
            current_day = parse_iso_day(event.start)
            if current_day == previous_day + timedelta(days=1):
                run.append(event)
            else:
                merged.extend(collapse_run(run))
                run = [event]
        merged.extend(collapse_run(run))

    return sorted(passthrough + merged, key=sort_key)


def sort_key(event: EventDraft) -> tuple[str, int, str]:
    return (event.start, 0 if not event.all_day else 1, slug(event.title))


def main() -> None:
    rows = rows_from_sheet(DEFAULT_SOURCE)
    events = merge_consecutive_single_day_events(sorted(build_events(rows), key=sort_key))

    category_counts: dict[str, int] = defaultdict(int)
    confirmation_count = 0
    for event in events:
        category_counts[event.category] += 1
        if event.needs_confirmation:
            confirmation_count += 1

    payload = {
        "meta": {
            "schoolYear": "2026/2027",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceWorkbook": str(DEFAULT_SOURCE),
            "totalEvents": len(events),
            "confirmationCount": confirmation_count,
            "categoryCounts": dict(sorted(category_counts.items())),
            "notes": [
                "Źródłem prawdy jest plik kalendarz_dzienny_role_2026_2027.xlsx.",
                "Wpis BS II z 2027-06-17 pozostaje oznaczony jako wymagający doprecyzowania.",
                "Wpis z 2027-06-28 pozostaje zachowany jako punkt do wyjaśnienia.",
            ],
        },
        "events": [event.to_dict(idx + 1) for idx, event in enumerate(events)],
    }

    DEFAULT_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved {DEFAULT_OUTPUT}")
    print(f"events: {len(events)}")


if __name__ == "__main__":
    main()
